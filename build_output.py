"""Build the publish-ready allocation workbook from the engine result."""
from __future__ import annotations
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

FONT = "Calibri"
INK = "1F2A33"
HEADER_FILL = "1F3B4D"      # deep slate
HEADER_FONT = "FFFFFF"
BAND = "F2F5F7"
ACCENT = "C9A227"           # muted gold
GREEN = "E4EFE6"; AMBER = "FBF3DD"; RED = "F7E3E3"
THIN = Side(style="thin", color="D6DCE1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=HEADER_FONT, size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.font = Font(name=FONT, bold=True, size=13, color=INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(2, 1, f"JKCL AP Pendency · generated {dt.date.today():%d-%b-%Y} · new allocation logic")
    ws.cell(2, 1).font = Font(name=FONT, italic=True, size=9, color="6B7780")


def _autofit(ws, df, start_row):
    for i, col in enumerate(df.columns, 1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col].head(200))) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 42)


def build(result: pd.DataFrame, out_path: str):
    wb = Workbook()

    # ---------- Allocation ----------
    ws = wb.active
    ws.title = "Allocation"
    ncol = len(result.columns)
    _title(ws, "Overall — JKCL", ncol)
    hr = 4
    ws.append([]); ws.append([])  # rows 2,3 spacer (title used 1-2)
    # write header at row 4
    for j, col in enumerate(result.columns, 1):
        ws.cell(hr, j, col)
    _hdr(ws, ncol, row=hr)
    review_idx = list(result.columns).index("Needs Review")
    for r_i, (_, row) in enumerate(result.iterrows(), start=hr + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(r_i, j, val)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.border = BORDER
            if r_i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=BAND)
        if str(row["Needs Review"]):
            ws.cell(r_i, review_idx + 1).fill = PatternFill("solid", fgColor=AMBER)
    ws.freeze_panes = f"A{hr+1}"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(ncol)}{hr+len(result)}"
    # widths
    base = {"DCN": 22, "Vendor Name": 30, "Vendor Group": 18, "Assigned To": 22,
            "Assigned Email": 30, "Claimed By": 26, "Needs Review": 24}
    for j, col in enumerate(result.columns, 1):
        ws.column_dimensions[get_column_letter(j)].width = base.get(col, 13)
    alloc_rows = len(result)
    alloc_last = hr + alloc_rows
    team_col = get_column_letter(list(result.columns).index("Team") + 1)
    role_col = get_column_letter(list(result.columns).index("Role") + 1)
    name_col = get_column_letter(list(result.columns).index("Assigned To") + 1)
    util_col = get_column_letter(list(result.columns).index("Utility") + 1)
    doc_col = get_column_letter(list(result.columns).index("Doc Type") + 1)
    amt_col = get_column_letter(list(result.columns).index("Amount") + 1)
    rng = lambda c: f"Allocation!${c}${hr+1}:${c}${alloc_last}"

    # ---------- Overall (supervisor-wise pivot, dynamic COUNTIFS) ----------
    ov = wb.create_sheet("Overall")
    people = (result.groupby(["Team", "Assigned To"]).size()
              .reset_index().sort_values(["Team", "Assigned To"]))
    _title(ov, "Overall — Supervisor-wise Allocation", 5)
    head = ["Team", "Assigned To", "Total", "Parker", "Poster"]
    for j, h in enumerate(head, 1):
        ov.cell(4, j, h)
    _hdr(ov, len(head), row=4)
    r = 5
    role_ct = (result.groupby(["Team", "Assigned To", "Role"]).size()
               .unstack(fill_value=0))
    for _, prow in people.iterrows():
        t, n = prow["Team"], prow["Assigned To"]
        parker = int(role_ct.loc[(t, n), "Parker"]) if ("Parker" in role_ct.columns and (t, n) in role_ct.index) else 0
        poster = int(role_ct.loc[(t, n), "Poster"]) if ("Poster" in role_ct.columns and (t, n) in role_ct.index) else 0
        ov.cell(r, 1, t); ov.cell(r, 2, n)
        ov.cell(r, 3, parker + poster)
        ov.cell(r, 4, parker)
        ov.cell(r, 5, poster)
        for c in range(1, 6):
            cell = ov.cell(r, c); cell.font = Font(name=FONT, size=9, color=INK); cell.border = BORDER
            if r % 2 == 0: cell.fill = PatternFill("solid", fgColor=BAND)
        r += 1
    tot = r
    ov.cell(tot, 2, "Grand Total").font = Font(name=FONT, bold=True, color=INK)
    gt_total = int(len(result))
    gt_parker = int((result["Role"] == "Parker").sum())
    gt_poster = int((result["Role"] == "Poster").sum())
    for c, val in [(3, gt_total), (4, gt_parker), (5, gt_poster)]:
        ov.cell(tot, c, val)
        ov.cell(tot, c).font = Font(name=FONT, bold=True, color=INK)
        ov.cell(tot, c).fill = PatternFill("solid", fgColor=GREEN)
    for w, c in zip([16, 24, 10, 10, 10], "ABCDE"):
        ov.column_dimensions[c].width = w
    ov.freeze_panes = "A5"

    # ---------- AP (team / type summary) ----------
    ap = wb.create_sheet("AP")
    _title(ap, "AP — Pendency Summary", 4)
    ap.cell(4, 1, "Team"); ap.cell(4, 2, "Total"); ap.cell(4, 3, "PO"); ap.cell(4, 4, "Non-PO")
    _hdr(ap, 4, row=4)
    teams = sorted(result["Team"].unique())
    by_team_total = result["Team"].value_counts()
    po_ct = result[result["Doc Type"] == "PO"]["Team"].value_counts()
    npo_ct = result[result["Doc Type"] == "NON PO"]["Team"].value_counts()
    r = 5
    for t in teams:
        ap.cell(r, 1, t)
        ap.cell(r, 2, int(by_team_total.get(t, 0)))
        ap.cell(r, 3, int(po_ct.get(t, 0)))
        ap.cell(r, 4, int(npo_ct.get(t, 0)))
        for c in range(1, 5):
            cell = ap.cell(r, c); cell.font = Font(name=FONT, size=9, color=INK); cell.border = BORDER
            if r % 2 == 0: cell.fill = PatternFill("solid", fgColor=BAND)
        r += 1
    ap.cell(r, 1, "Grand Total").font = Font(name=FONT, bold=True)
    for c, val in [(2, int(len(result))),
                   (3, int((result["Doc Type"] == "PO").sum())),
                   (4, int((result["Doc Type"] == "NON PO").sum()))]:
        ap.cell(r, c, val)
        ap.cell(r, c).font = Font(name=FONT, bold=True)
        ap.cell(r, c).fill = PatternFill("solid", fgColor=GREEN)
    # utility split block
    r2 = r + 2
    ap.cell(r2, 1, "Utility").font = Font(name=FONT, bold=True, color=INK)
    ap.cell(r2, 2, "Count").font = Font(name=FONT, bold=True, color=INK)
    _hdr(ap, 2, row=r2)
    util_ct = result["Utility"].value_counts()
    for k, u in enumerate(["Utility", "Non Utility", "Unmapped"], 1):
        ap.cell(r2 + k, 1, u)
        ap.cell(r2 + k, 2, int(util_ct.get(u, 0)))
        for c in (1, 2):
            ap.cell(r2 + k, c).font = Font(name=FONT, size=9, color=INK); ap.cell(r2 + k, c).border = BORDER
    for w, c in zip([18, 12, 10, 10], "ABCD"):
        ap.column_dimensions[c].width = w

    # ---------- Review ----------
    rv = wb.create_sheet("Review")
    flagged = result[result["Needs Review"] != ""].copy()
    _title(rv, f"Review queue — {len(flagged)} rows need a human check before publishing", len(result.columns))
    for j, col in enumerate(result.columns, 1):
        rv.cell(4, j, col)
    _hdr(rv, len(result.columns), row=4)
    for r_i, (_, row) in enumerate(flagged.iterrows(), start=5):
        for j, val in enumerate(row, 1):
            cell = rv.cell(r_i, j, val); cell.font = Font(name=FONT, size=9, color=INK); cell.border = BORDER
            if r_i % 2 == 0: cell.fill = PatternFill("solid", fgColor=BAND)
        rv.cell(r_i, review_idx + 1).fill = PatternFill("solid", fgColor=AMBER)
    rv.freeze_panes = "A5"
    if len(flagged):
        rv.auto_filter.ref = f"A4:{get_column_letter(len(result.columns))}{4+len(flagged)}"
    for j, col in enumerate(result.columns, 1):
        rv.column_dimensions[get_column_letter(j)].width = base.get(col, 13)

    wb.save(out_path)
    return out_path
